"""Utility functions for Linear import."""

import csv
import os
from datetime import datetime
import re

MAX_PROJECT_NAME_LENGTH = 80
MAX_ISSUE_TITLE_LENGTH = 255


_KNOWN_HTML_TAGS = frozenset([
    'a', 'b', 'br', 'code', 'del', 'div', 'em', 'h1', 'h2', 'h3', 'h4',
    'h5', 'h6', 'hr', 'i', 'img', 'ins', 'li', 'ol', 'p', 'pre', 's',
    'span', 'strong', 'sub', 'sup', 'table', 'tbody', 'td', 'th', 'thead',
    'tr', 'u', 'ul',
])


def strip_html_tags(text: str) -> str:
    """Remove HTML tags from text while preserving non-HTML angle-bracket content.

    Strips known HTML tags (``<span …>``, ``<p>``, ``<div>`` etc.) and any
    tag whose attributes contain ``=`` (e.g. ``<span data-foo="bar">``).
    Leaves Markdown auto-links (``<https://…>``) and plain angle-bracket
    content (``<total entries>``) untouched.
    """
    if not text:
        return text

    def _replace_opening(m):
        tag_name = m.group(1).lower()
        attrs = m.group(2)
        if tag_name in _KNOWN_HTML_TAGS:
            return ''
        if '=' in attrs:
            return ''
        return m.group(0)

    text = re.sub(
        r'<(?!https?://)([a-zA-Z][a-zA-Z0-9]*)(\b[^>]*)/?>',
        _replace_opening,
        text,
    )
    text = re.sub(r'</([a-zA-Z][a-zA-Z0-9]*)\s*>', lambda m: '' if m.group(1).lower() in _KNOWN_HTML_TAGS else m.group(0), text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text


def parse_datetime(dt_str: str) -> str:
    """Parse a date/time string to ISO 8601 format (``YYYY-MM-DDTHH:MM:SSZ``).

    Handles common short formats found in ProductBoard exports such as
    ``4/3/25 5:14`` or ``3/23/2026 16:15``.  Returns *None* if unparseable.
    """
    if not dt_str or not dt_str.strip():
        return None

    dt_str = dt_str.strip().strip('"')

    formats = [
        "%m/%d/%y %H:%M",       # 4/3/25 5:14
        "%m/%d/%Y %H:%M",       # 3/23/2026 16:15
        "%m/%d/%y %I:%M %p",    # 4/3/25 5:14 PM
        "%m/%d/%Y %I:%M %p",    # 3/23/2026 5:14 PM
        "%Y-%m-%d %H:%M:%S",    # 2026-03-23 16:15:00
        "%Y-%m-%dT%H:%M:%S",    # 2026-03-23T16:15:00
        "%Y-%m-%dT%H:%M:%SZ",   # 2026-03-23T16:15:00Z
        "%m/%d/%y",              # 4/3/25
        "%m/%d/%Y",              # 4/3/2025
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(dt_str, fmt)
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue

    return None


def truncate_name(name: str, max_length: int = MAX_PROJECT_NAME_LENGTH) -> str:
    """Truncate a name to max length, adding ellipsis if needed."""
    if len(name) > max_length:
        return name[:max_length - 3] + "..."
    return name


def parse_date(date_str: str) -> str:
    """Parse various date formats to ISO format (YYYY-MM-DD)."""
    if not date_str or not date_str.strip():
        return None
    
    date_str = date_str.strip().strip('"')

    # Skip known non-date keywords
    lower = date_str.lower()
    if lower in ("tbd", "done", "not needed", "na", "n/a", "no ux needed",
                  "not needed", "no term sheet needed"):
        return None
    if lower.startswith(("❌", "🚀 shipped", "🚢 ship")):
        return None

    # Strip leading day-of-week prefix (e.g. "Mon ", "Tue ")
    date_str = re.sub(r'^(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+', '', date_str, flags=re.IGNORECASE)

    # Strip leading "Before " prefix (e.g. "Before Nov 2025")
    date_str = re.sub(r'^Before\s+', '', date_str, flags=re.IGNORECASE)
    
    # Try various formats
    formats = [
        "%m/%d/%Y",      # 1/5/2026
        "%m/%d/%y",      # 1/5/26
        "%Y-%m-%d",      # 2026-01-05
        "%Y/%m/%d",      # 2026/01/05
        "%d/%m/%Y",      # 05/01/2026 (European)
        "%d-%m-%Y",      # 05-01-2026
        "%b %d, %Y",    # Mar 06, 2026
        "%B %d, %Y",    # March 06, 2026
        "%b %d %Y",     # Mar 06 2026
        "%B %d %Y",     # March 06 2026
    ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    # Formats without a year — assume current year
    no_year_formats = [
        "%b %d",         # Mar 6
        "%B %d",         # March 6
        "%m/%d",         # 3/6
    ]
    current_year = datetime.now().year
    for fmt in no_year_formats:
        try:
            dt = datetime.strptime(date_str, fmt).replace(year=current_year)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Month + year (e.g. "February 2026", "Feb 2026") → first of month
    month_year_formats = [
        "%B %Y",     # February 2026
        "%b %Y",     # Feb 2026
        "%B'%Y",     # February'2026
        "%b'%Y",     # Feb'2026
    ]
    for fmt in month_year_formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Quarter notation (e.g. "Q2 2026", "Q3 2025") → first day of quarter
    quarter_match = re.match(r'^Q([1-4])\s*[\'"]?(\d{4})$', date_str, re.IGNORECASE)
    if quarter_match:
        q, year = int(quarter_match.group(1)), int(quarter_match.group(2))
        month = {1: 1, 2: 4, 3: 7, 4: 10}[q]
        return f"{year}-{month:02d}-01"

    # Half-year notation (e.g. "H1 2026", "H2 2025") → first day of half
    half_match = re.match(r'^H([12])\s*[\'"]?(\d{4})$', date_str, re.IGNORECASE)
    if half_match:
        h, year = int(half_match.group(1)), int(half_match.group(2))
        month = 1 if h == 1 else 7
        return f"{year}-{month:02d}-01"

    return None


def parse_last_date(raw_str: str) -> str:
    """Parse a potentially comma-separated multi-date string, returning the last parseable date.

    Handles quoted dates (``"Mon Mar 23, 2026"``), non-date keywords
    (``Done``, ``TBD``), and mixed formats like::

        Before Nov 2025, Done, "Mon Jan 19, 2026"
        "Mon Jan 12, 2026", "Mon Mar 2, 2026", Done
    """
    if not raw_str or not raw_str.strip():
        return None

    raw_str = raw_str.strip()

    # Fast path: no comma and no quote → single value
    if ',' not in raw_str and '"' not in raw_str:
        return parse_date(raw_str)

    # Tokenize respecting quoted strings (which may contain commas)
    tokens = []
    i = 0
    while i < len(raw_str):
        while i < len(raw_str) and raw_str[i] in (' ', ','):
            i += 1
        if i >= len(raw_str):
            break

        if raw_str[i] == '"':
            end = raw_str.find('"', i + 1)
            if end == -1:
                tokens.append(raw_str[i + 1:].strip())
                break
            tokens.append(raw_str[i + 1:end].strip())
            i = end + 1
        else:
            end = i
            while end < len(raw_str) and raw_str[end] not in (',', '"'):
                end += 1
            token = raw_str[i:end].strip()
            if token:
                tokens.append(token)
            i = end

    last_date = None
    for token in tokens:
        parsed = parse_date(token)
        if parsed:
            last_date = parsed

    return last_date


def normalize_status(status: str, status_map: dict) -> str:
    """Normalize a status value using the provided mapping."""
    if not status:
        return None
    
    status = status.strip()
    
    # Try exact match first
    if status in status_map:
        return status_map[status]
    
    # Try case-insensitive match
    status_lower = status.lower()
    for key, value in status_map.items():
        if key.lower() == status_lower:
            return value
    
    return None


def normalize_priority(priority: str, priority_map: dict) -> int:
    """Normalize a priority value using the provided mapping."""
    if not priority:
        return 0
    
    priority = priority.strip()
    
    # Try exact match first
    if priority in priority_map:
        return priority_map[priority]
    
    # Try case-insensitive match
    priority_lower = priority.lower()
    for key, value in priority_map.items():
        if key.lower() == priority_lower:
            return value
    
    return 0


def extract_project_name_from_filename(filename: str) -> str:
    """Extract project name from a CSV filename."""
    # Remove path
    name = filename.split("/")[-1].split("\\")[-1]
    
    # Remove extension
    name = re.sub(r"\.(csv|tsv)$", "", name, flags=re.IGNORECASE)
    
    prefixes = [
        r"^Copy of\s+",
        r"^[A-Z]{2,5}\s*-\s*Project Tracker\s+H\d+\s*\d*\s*-\s*",
        r"^Project Tracker\s*-\s*",
    ]
    for prefix in prefixes:
        name = re.sub(prefix, "", name, flags=re.IGNORECASE)
    
    return name.strip()


def parse_estimate(value: str) -> float:
    """Parse estimate value (story points or effort days)."""
    if not value:
        return None
    
    try:
        return float(value.strip())
    except (ValueError, AttributeError):
        return None


def priority_from_ranking(ranking_str: str, priority_ranges: list, default: int = 0) -> int:
    """Convert a numeric ranking value to Linear priority (0-4) using range buckets.
    
    priority_ranges is a list of dicts with 'max' and 'priority' keys:
      [{"max": 100, "priority": 1}, {"max": 200, "priority": 2}, ...]
    
    Returns the priority for the first range where ranking <= max.
    """
    if not ranking_str:
        return default
    
    try:
        ranking = float(ranking_str.strip())
    except (ValueError, AttributeError):
        return default
    
    for bucket in priority_ranges:
        if ranking <= bucket.get("max", float("inf")):
            return bucket.get("priority", default)
    
    return default


def convert_xlsx_to_csv(xlsx_path: str, sheet_name: str = None) -> str:
    """Convert an Excel (.xlsx) file to CSV, returning the path to the new CSV.

    Requires the ``openpyxl`` package (``pip install openpyxl``).
    Converts the sheet named *sheet_name*, or the active sheet if not given.
    The output CSV is written next to the source file with a ``.csv``
    extension (with sheet name suffix when a non-default sheet is selected).
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Excel support requires openpyxl. "
            "Install with: pip install openpyxl"
        )

    base = os.path.splitext(xlsx_path)[0]
    if sheet_name:
        safe_name = re.sub(r'[^\w\s-]', '', sheet_name).strip().replace(' ', '_')
        csv_path = f"{base}_{safe_name}.csv"
    else:
        csv_path = base + ".csv"

    print(f"  Converting {os.path.basename(xlsx_path)} to CSV...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]
        print(f"  Using sheet: {sheet_name}")
    else:
        ws = wb.active

    headers = []
    for cell in ws[1]:
        val = cell.value
        if val is not None:
            header = " ".join(str(val).split())
            headers.append(header)
        else:
            headers.append("")

    # Trim trailing empty headers
    while headers and not headers[-1]:
        headers.pop()

    num_cols = len(headers)
    row_count = 0

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in ws.iter_rows(min_row=2, max_col=num_cols, values_only=True):
            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {}
            for col_idx, header in enumerate(headers):
                cell = row[col_idx] if col_idx < len(row) else None
                row_data[header] = str(cell) if cell is not None else ""
            writer.writerow(row_data)
            row_count += 1

    print(f"  Converted {row_count} rows -> {os.path.basename(csv_path)}")
    return csv_path


def convert_numbers_to_csv(numbers_path: str) -> str:
    """Convert an Apple Numbers file to CSV, returning the path to the new CSV.
    
    Requires the ``numbers-parser`` package (``pip install numbers-parser``).
    Converts the first table in the first sheet.  The output CSV is written
    next to the source file with a ``.csv`` extension.
    """
    try:
        from numbers_parser import Document
    except ImportError:
        raise ImportError(
            "Apple Numbers support requires numbers-parser. "
            "Install with: pip install numbers-parser"
        )

    csv_path = os.path.splitext(numbers_path)[0] + ".csv"

    print(f"  Converting {os.path.basename(numbers_path)} to CSV...")
    doc = Document(numbers_path)
    table = doc.sheets[0].tables[0]

    headers = []
    for col in range(table.num_cols):
        cell = table.cell(0, col)
        headers.append(str(cell.value) if cell.value is not None else "")

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row_idx in range(1, table.num_rows):
            row_data = {}
            for col_idx, header in enumerate(headers):
                cell = table.cell(row_idx, col_idx)
                row_data[header] = str(cell.value) if cell.value is not None else ""
            writer.writerow(row_data)

    print(f"  Converted {table.num_rows - 1} rows -> {os.path.basename(csv_path)}")
    return csv_path
